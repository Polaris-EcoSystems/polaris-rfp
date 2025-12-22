from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from ..observability.logging import get_logger
from ..services import content_repo
from ..services.contracting_repo import (
    add_budget_version,
    add_contract_doc_version,
    get_case_by_id,
    get_contract_template,
    get_contract_template_version,
)
from ..repositories.rfp.proposals_repo import get_proposal_by_id
from ..repositories.rfp.rfps_repo import get_rfp_by_id
from ..services.s3_assets import get_object_bytes, put_object_bytes
from ..services.contracting_schemas import ContractingKeyTerms


log = get_logger("contracting_docgen")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _safe_str(v: Any) -> str:
    return str(v or "").strip()


def _contract_output_key(case_id: str, kind: str, ext: str) -> str:
    # Store generated artifacts under a stable namespace.
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"contracting/{_safe_str(case_id)}/{kind}/{ts}_{ext}".replace("//", "/")


def render_contract_docx(
    *,
    case_id: str,
    template_id: str,
    template_version_id: str | None,
    render_inputs: dict[str, Any] | None,
    created_by_user_sub: str | None,
) -> dict[str, Any]:
    """
    Render a DOCX contract from a template stored in S3 (docxtpl).
    Persists the rendered artifact to S3 and records a ContractDocumentVersion.
    """
    cid = _safe_str(case_id)
    if not cid:
        raise ValueError("case_id is required")
    tid = _safe_str(template_id)
    if not tid:
        raise ValueError("template_id is required")

    case = get_case_by_id(cid)
    if not case:
        raise ValueError("Contracting case not found")

    proposal_id = _safe_str(case.get("proposalId"))
    if not proposal_id:
        raise ValueError("Case missing proposalId")

    proposal = get_proposal_by_id(proposal_id, include_sections=True)
    if not proposal:
        raise ValueError("Proposal not found")

    rfp = get_rfp_by_id(_safe_str(case.get("rfpId"))) or {}
    company = None
    company_id = _safe_str(case.get("companyId") or proposal.get("companyId"))
    if company_id:
        company = content_repo.get_company_by_company_id(company_id)

    tpl = get_contract_template(tid)
    if not tpl:
        raise ValueError("Contract template not found")

    version_id = _safe_str(template_version_id) or _safe_str(tpl.get("currentVersionId"))
    if not version_id:
        raise ValueError("Contract template has no currentVersionId; upload a template version first")

    tpl_ver = get_contract_template_version(tid, version_id)
    if not tpl_ver:
        raise ValueError("Contract template version not found")
    s3_key = _safe_str(tpl_ver.get("s3Key"))
    if not s3_key:
        raise ValueError("Template version missing s3Key")

    try:
        from docxtpl import DocxTemplate
    except Exception as e:
        raise RuntimeError("DOCX template dependency not installed (docxtpl)") from e

    template_bytes = get_object_bytes(key=s3_key, max_bytes=20 * 1024 * 1024)
    if not template_bytes:
        raise RuntimeError("Template object is empty or missing")

    # docxtpl can load from a file-like object.
    doc = DocxTemplate(io.BytesIO(template_bytes))

    # Build render context.
    # Validate key terms at generation time (fail fast).
    kt_obj = case.get("keyTerms") if isinstance(case.get("keyTerms"), dict) else {}
    try:
        kt_norm = ContractingKeyTerms.model_validate(kt_obj).model_dump(mode="json")
    except Exception as e:
        raise ValueError(f"Invalid key terms: {str(e) or 'invalid'}") from e

    context = {
        "case": case,
        "keyTerms": kt_norm,
        "proposal": proposal,
        "rfp": rfp,
        "company": company or {},
        "renderInputs": render_inputs if isinstance(render_inputs, dict) else {},
        "generatedAt": _now_iso(),
    }
    # Allow callers to override/extend the base context.
    if isinstance(render_inputs, dict):
        context.update({k: v for k, v in render_inputs.items() if k not in ("case", "proposal", "rfp", "company")})

    try:
        doc.render(context)
    except Exception as e:
        log.exception("contract_docx_render_failed", caseId=cid, templateId=tid, templateVersionId=version_id)
        raise RuntimeError(f"Failed to render DOCX template: {str(e) or 'render_failed'}") from e

    buf = io.BytesIO()
    doc.save(buf)
    out_bytes = buf.getvalue()
    if not out_bytes:
        raise RuntimeError("Rendered document was empty")

    out_key = _contract_output_key(cid, "contract", "contract.docx")
    put_object_bytes(
        key=out_key,
        data=out_bytes,
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

    version = add_contract_doc_version(
        case_id=cid,
        source_template_id=tid,
        source_template_version_id=version_id,
        render_inputs=render_inputs if isinstance(render_inputs, dict) else {},
        docx_s3_key=out_key,
        pdf_s3_key=None,
        created_by_user_sub=created_by_user_sub,
    )
    return {"version": version, "docxS3Key": out_key}


def generate_budget_xlsx(
    *,
    case_id: str,
    budget_model: dict[str, Any] | None,
    created_by_user_sub: str | None,
) -> dict[str, Any]:
    """
    Generate an internal budget workbook (.xlsx) from a JSON model.
    Persists the workbook to S3 and records a BudgetVersion.
    """
    cid = _safe_str(case_id)
    if not cid:
        raise ValueError("case_id is required")

    case = get_case_by_id(cid)
    if not case:
        raise ValueError("Contracting case not found")

    # Default: if caller didn't provide a model, try to derive from proposal.budgetBreakdown.
    model = budget_model if isinstance(budget_model, dict) else {}
    if not model:
        proposal = get_proposal_by_id(_safe_str(case.get("proposalId")), include_sections=False) or {}
        bb = proposal.get("budgetBreakdown")
        if isinstance(bb, dict):
            model = bb
        else:
            model = {"items": [], "notes": "", "currency": "USD"}

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:
        raise RuntimeError("XLSX dependency not installed (openpyxl)") from e

    raw_items = model.get("items")
    items_in: list[Any] = raw_items if isinstance(raw_items, list) else []
    items: list[dict[str, Any]] = [x for x in items_in if isinstance(x, dict)]

    # Normalize + compute totals.
    def _num(v: Any) -> float:
        try:
            if v is None or v == "":
                return 0.0
            return float(v)
        except Exception:
            return 0.0

    norm_items: list[dict[str, Any]] = []
    for it in items:
        role = _safe_str(it.get("role"))
        phase = _safe_str(it.get("phase"))
        name = _safe_str(it.get("name")) or role or "Line item"
        rate = _num(it.get("rate"))
        hours = _num(it.get("hours"))
        qty = _num(it.get("qty")) or 0.0
        if hours <= 0 and qty > 0:
            hours = qty
        cost = _num(it.get("cost"))
        if cost <= 0 and rate > 0 and hours > 0:
            cost = rate * hours
        norm_items.append(
            {
                "phase": phase,
                "name": name,
                "role": role,
                "rate": rate,
                "hours": hours,
                "cost": cost,
                "notes": _safe_str(it.get("notes")),
            }
        )

    total = sum(float(x.get("cost") or 0.0) for x in norm_items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    label_font = Font(bold=True)

    ws["A1"] = "Project Budget (Internal)"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Generated at"
    ws["B3"] = _now_iso()
    ws["A4"] = "Total"
    ws["B4"] = total
    ws["A3"].font = label_font
    ws["A4"].font = label_font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # Line items sheet
    ws2 = wb.create_sheet("Line Items")
    headers = ["Phase", "Name", "Role", "Rate", "Hours", "Cost", "Notes"]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for it in norm_items:
        ws2.append(
            [
                it.get("phase") or "",
                it.get("name") or "",
                it.get("role") or "",
                float(it.get("rate") or 0.0),
                float(it.get("hours") or 0.0),
                float(it.get("cost") or 0.0),
                it.get("notes") or "",
            ]
        )

    # Basic widths
    widths = [18, 30, 22, 12, 12, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(ord("A") + i - 1)].width = w

    # Totals row
    ws2.append(["", "", "", "", "Total", total, ""])
    total_row = ws2.max_row
    ws2.cell(row=total_row, column=5).font = label_font
    ws2.cell(row=total_row, column=6).font = label_font

    # Notes sheet
    ws3 = wb.create_sheet("Assumptions & Notes")
    ws3["A1"] = "Assumptions & Notes"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A3"] = str(model.get("notes") or "")
    ws3.column_dimensions["A"].width = 100
    ws3["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    out_key = _contract_output_key(cid, "budget", "budget.xlsx")
    put_object_bytes(
        key=out_key,
        data=data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    version = add_budget_version(
        case_id=cid,
        budget_model=model,
        xlsx_s3_key=out_key,
        created_by_user_sub=created_by_user_sub,
    )
    return {"version": version, "xlsxS3Key": out_key, "total": total}

